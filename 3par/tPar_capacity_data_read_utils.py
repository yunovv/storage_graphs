import openpyxl
import sys

# Read Excel data
def get_arrays_xlsx(sheet_name):
    blacklist = get_blacklist()
    print('Read Excel data from:',sheet_name)
    print('searching for \'X%\' \'Y%\' and \'System Name\'')

    found_points_xy = []

    try:
        input_wb = openpyxl.load_workbook(sheet_name)
    except:
        print('Can\'t read from',sheet_name)
        sys.exit(1)
    input_ws = input_wb.worksheets[0]

    headers_row = -1
    sysname_index = -1
    x_index = -1
    y_index = -1
    for r,row in enumerate(input_ws.iter_rows()):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if 'System Name' in whole_row and 'X%' in whole_row and 'Y%' in whole_row:
            headers_row = r
            for i in range(0,len(whole_row)):
                if whole_row[i] == 'X%':
                    x_index = i
                elif whole_row[i] == 'Y%':
                    y_index = i
                elif whole_row[i] == 'System Name':
                    sysname_index = i
            break
    for r,row in enumerate(input_ws.iter_rows(min_row=headers_row+2)):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if whole_row[sysname_index] != None and whole_row[x_index] != None and whole_row[y_index] != None:
            if whole_row[sysname_index] in blacklist:
                print('blacklisted:',whole_row[sysname_index])
            else:
                found_points_xy.append([float(whole_row[x_index]),float(whole_row[y_index]),str(whole_row[sysname_index])])

    print('Found systems:',len(found_points_xy),found_points_xy)
    return found_points_xy

def get_blacklist():
    #print('Checking blacklist')
    blacklist = []
    try:
        blacklist_file = open("blacklist")
        for line in blacklist_file:
            if line != '\n':
                for word in line.split(','):
                    blacklist.append(word.strip())

    except IOError:
        print("File \'blacklist\' not accessible")
        return blacklist

    blacklist_file.close()
    print('blacklist:',blacklist)

    return blacklist

def get_arrays_xlsx_full(sheet_name):
    blacklist = get_blacklist()
    print('Read Excel data from:',sheet_name)
    print('searching for \'X%\' \'Y%\' and \'System Name\'')

    found_arrays = []

    try:
        input_wb = openpyxl.load_workbook(sheet_name)
    except:
        print('Can\'t read from',sheet_name)
        sys.exit(1)
    input_ws = input_wb.worksheets[0]

    headers_row = -1
    sysname_index = -1
    x_index = -1
    y_index = -1
    raw_index = -1
    alloc_index = -1
    for r,row in enumerate(input_ws.iter_rows()):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if 'System Name' in whole_row and 'X%' in whole_row and 'Y%' in whole_row and 'Total RAW' in whole_row and 'Allocated RAW' in whole_row:
            headers_row = r
            for i in range(0,len(whole_row)):
                if whole_row[i] == 'System Name':
                    sysname_index = i
                elif whole_row[i] == 'X%':
                    x_index = i
                elif whole_row[i] == 'Y%':
                    y_index = i
                elif whole_row[i] == 'Total RAW':
                    raw_index = i
                elif whole_row[i] == 'Allocated RAW':
                    alloc_index = i
            break

    if headers_row < 0:
        print('Can\'t find headers X%   Y%   System Name   Total RAW   Allocated RAW')
        return found_arrays


    for r,row in enumerate(input_ws.iter_rows(min_row=headers_row+2)):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if whole_row[sysname_index] != None and whole_row[x_index] != None and whole_row[y_index] != None and whole_row[raw_index] != None and whole_row[alloc_index] != None:
            if whole_row[sysname_index] in blacklist:
                print('blacklisted:',whole_row[sysname_index])
            else:
                found_arrays.append([str(whole_row[sysname_index]),float(whole_row[x_index]),float(whole_row[y_index]),float(whole_row[raw_index]),float(whole_row[alloc_index])])

    print('Found systems:',len(found_arrays),found_arrays)
    return found_arrays
