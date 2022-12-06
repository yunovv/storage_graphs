from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import openpyxl
import os
import re
from datetime import date, datetime


output_directory= os.getcwd()

now = datetime.now()

# Creare output Workbook
inventory_wb = Workbook()
default_sheet = inventory_wb.get_sheet_by_name('Sheet')
inventory_wb.remove_sheet(default_sheet)

# Define borders and styles
allsides_border_th = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
right_border_th = Border(right=Side(style='thin'))
botton_border_th = Border(bottom=Side(style='thin'))
corner_border_th = Border(right=Side(style='thin'), bottom=Side(style='thin'))

greyFill = PatternFill(start_color='D0D0D0',end_color='D0D0D0',fill_type='solid')



def write_collected_data_to_worksheet(ws_title,ws_header,data):

    ws = inventory_wb.create_sheet()
    ws.title = ws_title
    rownum = 0

    if ws_header != None:
        create_header(ws,ws_header)
        rownum = 1


    for instance in data:
        if len(instance) > 0:
            for component in instance:
                letter = 'Z'
                rownum += 1
                for param in component:
                    letter = get_nextletter(letter)
                    ws[letter + str(rownum)] = prepare_string(param)
                make_side_border(ws,letter + str(rownum))
            make_underscore_border(ws,letter,rownum)

    return

def create_header(ws,ws_header):
    letter = 'Z'
    for columnt_name in ws_header:
        letter = get_nextletter(letter)
        ws[letter + '1'] = columnt_name
        make_allside_borders(ws, letter + '1')
        paint_cell_grey(ws, letter + '1')

    # Make Filters and Pin to view
    ws.auto_filter.ref = 'A1:' + str(letter + '1')
    ws.freeze_panes = "A2"

    return

def make_allside_borders(ws,cell_name):
    ws[str(cell_name)].border = allsides_border_th
    return

def make_side_border(ws,cell_name):
    ws[str(cell_name)].border = right_border_th
    return

def make_underscore_border(ws,end_letter,rownum):
    for row in ws['A'+str(rownum) : end_letter + str(rownum)]:
        for cell in row:
            cell.border = botton_border_th
    ws[end_letter + str(rownum)].border = corner_border_th
    return

def paint_cell_grey(ws,cell_name):
    ws[str(cell_name)].fill = greyFill
    return

def auto_allign_cells_in_workbook():
    for ws_name in inventory_wb.get_sheet_names():
        ws = inventory_wb.get_sheet_by_name(ws_name)
        for column_cells in ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 2
    return

def save_workbook(wb_name):
    if len(inventory_wb.worksheets) == 0:
        return

    auto_allign_cells_in_workbook()
    inventory_wb.save(wb_name + '.xlsx')
    print('Results saved to ' + wb_name)
    return

def get_nextletter(current):
    return 'A' if current == 'Z' else chr(ord(current) + 1)

def as_text(value):
    if value is None:
        return ""
    return str(value)

def prepare_string(input):
    find_illegal = re.findall(r'([\000-\010]|[\013-\014]|[\016-\037])',str(input))
    for bad in find_illegal:
        input = str(input).replace(bad,'')
    result = input
    return result
