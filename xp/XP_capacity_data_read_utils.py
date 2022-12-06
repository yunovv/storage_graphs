import openpyxl
import sys

# Read Excel data
def get_arrays_xlsx(sheet_name):
    blacklist = get_blacklist()
    print('Read Excel data from:',sheet_name)
    print('searching for \'total virtual capacity\' \'array capacity total\' \'total used capacity\' \'Pool\' and \'array name\'')

    found_points_xy = []


    try:
        input_wb = openpyxl.load_workbook(sheet_name)
    except:
        print('Can\'t read from',sheet_name)
        sys.exit(1)
    input_ws = input_wb.worksheets[0]

    headers_row = -1
    sysname_index = -1
    arr_cap_tot_index = -1
    tot_virt_cap_index = -1
    tot_used_cap_index = -1
    pool_index = -1
    for r,row in enumerate(input_ws.iter_rows()):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if 'total virtual capacity' in whole_row and 'array capacity total' in whole_row and 'total used capacity' in whole_row and 'array name' in whole_row and 'Pool' in whole_row:
            headers_row = r
            for i in range(0,len(whole_row)):
                if whole_row[i] == 'total virtual capacity':
                    tot_virt_cap_index = i
                elif whole_row[i] == 'array capacity total':
                    arr_cap_tot_index = i
                elif whole_row[i] == 'total used capacity':
                    tot_used_cap_index = i
                elif whole_row[i] == 'array name':
                    sysname_index = i
                elif whole_row[i] == 'Pool':
                    pool_index = i
            break
    for r,row in enumerate(input_ws.iter_rows(min_row=headers_row+2)):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if whole_row[sysname_index] != None and whole_row[arr_cap_tot_index] != None and whole_row[tot_virt_cap_index] != None and whole_row[tot_used_cap_index] != None and whole_row[pool_index] != None:
            overprov = whole_row[tot_virt_cap_index] / whole_row[arr_cap_tot_index] * 100
            util = whole_row[tot_used_cap_index] / whole_row[arr_cap_tot_index] * 100
            if whole_row[sysname_index] in blacklist:# or util<75:
                print('blacklisted:',whole_row[sysname_index])
            else:
                found_points_xy.append([float(util),float(overprov),str(whole_row[sysname_index])+'-'+str(whole_row[pool_index])])

    print('Found systems:',len(found_points_xy))
    return found_points_xy

def get_arrays_xlsx_compr(sheet_name):
    blacklist = get_blacklist()
    print('Read Excel data from:',sheet_name)
    print('searching for \'total virtual capacity\' \'array capacity total\' \'total used capacity\' \'compr\' \'Pool\' and \'array name\'')

    found_points_xy = []


    try:
        input_wb = openpyxl.load_workbook(sheet_name)
    except:
        print('Can\'t read from',sheet_name)
        sys.exit(1)
    input_ws = input_wb.worksheets[0]

    headers_row = -1
    sysname_index = -1
    arr_cap_tot_index = -1
    tot_virt_cap_index = -1
    tot_used_cap_index = -1
    pool_index = -1
    compr_index = -1
    for r,row in enumerate(input_ws.iter_rows()):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if 'total virtual capacity' in whole_row and 'array capacity total' in whole_row and 'total used capacity' in whole_row and 'array name' in whole_row and 'Pool' in whole_row and 'compr' in whole_row:
            headers_row = r
            for i in range(0,len(whole_row)):
                if whole_row[i] == 'total virtual capacity':
                    tot_virt_cap_index = i
                elif whole_row[i] == 'array capacity total':
                    arr_cap_tot_index = i
                elif whole_row[i] == 'total used capacity':
                    tot_used_cap_index = i
                elif whole_row[i] == 'array name':
                    sysname_index = i
                elif whole_row[i] == 'Pool':
                    pool_index = i
                elif whole_row[i] == 'compr':
                    compr_index = i
            break
    for r,row in enumerate(input_ws.iter_rows(min_row=headers_row+2)):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if whole_row[sysname_index] != None and whole_row[arr_cap_tot_index] != None and whole_row[tot_virt_cap_index] != None and whole_row[tot_used_cap_index] != None and whole_row[pool_index] != None:
            overprov = whole_row[tot_virt_cap_index] / whole_row[arr_cap_tot_index] * 100
            util = whole_row[tot_used_cap_index] / whole_row[arr_cap_tot_index] * 100
            if whole_row[sysname_index] in blacklist:
                print('blacklisted:',whole_row[sysname_index])
            else:
                if whole_row[compr_index] > 1:
                    found_points_xy.append([float(util),float(overprov),str(whole_row[sysname_index])+'-'+str(whole_row[pool_index])])

    print('Found systems:',len(found_points_xy))
    return found_points_xy

def get_arrays_xlsx_gr03(sheet_name):
    blacklist = get_blacklist()
    print('Read Excel data from:',sheet_name)
    print('searching for \'total virtual capacity\' \'array capacity total\' \'total used capacity\' \'compr\' \'Pool\' \'Capacity Expansion Ratio\' \'array physical capacity total\' \'array physical capacity remaining\' and \'array name\'')

    found_points_xy = []


    try:
        input_wb = openpyxl.load_workbook(sheet_name)
    except:
        print('Can\'t read from',sheet_name)
        sys.exit(1)
    input_ws = input_wb.worksheets[0]

    headers_row = -1
    sysname_index = -1
    arr_cap_tot_index = -1
    tot_virt_cap_index = -1
    tot_used_cap_index = -1
    pool_index = -1
    compr_index = -1
    expansion_index = -1
    arr_phy_cap_tot_index = -1
    arr_phy_cap_rem = -1
    for r,row in enumerate(input_ws.iter_rows()):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if 'total virtual capacity' in whole_row and 'array capacity total' in whole_row and 'total used capacity' in whole_row and 'array name' in whole_row and 'Pool' in whole_row and 'compr' in whole_row and 'Capacity Expansion Ratio' in whole_row and 'array physical capacity total' in whole_row and 'array physical capacity remaining' in whole_row:
            headers_row = r
            for i in range(0,len(whole_row)):
                if whole_row[i] == 'total virtual capacity':
                    tot_virt_cap_index = i
                elif whole_row[i] == 'array capacity total':
                    arr_cap_tot_index = i
                elif whole_row[i] == 'total used capacity':
                    tot_used_cap_index = i
                elif whole_row[i] == 'array name':
                    sysname_index = i
                elif whole_row[i] == 'Pool':
                    pool_index = i
                elif whole_row[i] == 'compr':
                    compr_index = i
                elif whole_row[i] == 'Capacity Expansion Ratio':
                    expansion_index = i
                elif whole_row[i] == 'array physical capacity total':
                    arr_phy_cap_tot_index = i
                elif whole_row[i] == 'array physical capacity remaining':
                    arr_phy_cap_rem = i
            break
    for r,row in enumerate(input_ws.iter_rows(min_row=headers_row+2)):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if whole_row[sysname_index] != None and whole_row[arr_cap_tot_index] != None and whole_row[tot_virt_cap_index] != None and whole_row[tot_used_cap_index] != None and whole_row[pool_index] != None:
            overprov = whole_row[tot_virt_cap_index] / whole_row[arr_cap_tot_index] * 100
            util = whole_row[tot_used_cap_index] / whole_row[arr_cap_tot_index] * 100
            V3 = whole_row[arr_phy_cap_tot_index]
            W3 = whole_row[arr_phy_cap_rem]
            L3 = whole_row[tot_virt_cap_index]
            N3 = whole_row[compr_index]
            point_x = (V3-W3)/V3*100
            point_y = (L3/(1+(N3-1)*0.7))/V3*100
            if whole_row[sysname_index] in blacklist:
                print('blacklisted:',whole_row[sysname_index])
            else:
                if (whole_row[compr_index] > 1 and util > 50 and overprov > 100) or whole_row[expansion_index] > whole_row[compr_index]:
                    found_points_xy.append([float(point_x),float(point_y),str(whole_row[sysname_index])+'-'+str(whole_row[pool_index])])

    print('Found systems:',len(found_points_xy))
    return found_points_xy

def get_arrays_xlsx_full(sheet_name):
    blacklist = get_blacklist()
    print('Read Excel data from:',sheet_name)
    print('searching for \'total virtual capacity\' \'array capacity total\' \'total used capacity\' \'compr\' \'Pool\' \'Capacity Expansion Ratio\' \'array physical capacity total\' \'array physical capacity remaining\' and \'array name\'')

    found_points_xy = []


    try:
        input_wb = openpyxl.load_workbook(sheet_name)
    except:
        print('Can\'t read from',sheet_name)
        sys.exit(1)
    input_ws = input_wb.worksheets[0]

    headers_row = -1
    sysname_index = -1
    arr_cap_tot_index = -1
    tot_virt_cap_index = -1
    tot_used_cap_index = -1
    pool_index = -1
    compr_index = -1
    expansion_index = -1
    arr_phy_cap_tot_index = -1
    arr_phy_cap_rem = -1
    for r,row in enumerate(input_ws.iter_rows()):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if 'total virtual capacity' in whole_row and 'array capacity total' in whole_row and 'total used capacity' in whole_row and 'array name' in whole_row and 'Pool' in whole_row and 'compr' in whole_row and 'Capacity Expansion Ratio' in whole_row and 'array physical capacity total' in whole_row and 'array physical capacity remaining' in whole_row:
            headers_row = r
            for i in range(0,len(whole_row)):
                if whole_row[i] == 'total virtual capacity':
                    tot_virt_cap_index = i
                elif whole_row[i] == 'array capacity total':
                    arr_cap_tot_index = i
                elif whole_row[i] == 'total used capacity':
                    tot_used_cap_index = i
                elif whole_row[i] == 'array name':
                    sysname_index = i
                elif whole_row[i] == 'Pool':
                    pool_index = i
                elif whole_row[i] == 'compr':
                    compr_index = i
                elif whole_row[i] == 'Capacity Expansion Ratio':
                    expansion_index = i
                elif whole_row[i] == 'array physical capacity total':
                    arr_phy_cap_tot_index = i
                elif whole_row[i] == 'array physical capacity remaining':
                    arr_phy_cap_rem = i
            break
    for r,row in enumerate(input_ws.iter_rows(min_row=headers_row+2)):
        whole_row = []
        for cell in row:
            whole_row.append(cell.value)
        if whole_row[sysname_index] != None and whole_row[arr_cap_tot_index] != None and whole_row[tot_virt_cap_index] != None and whole_row[tot_used_cap_index] != None and whole_row[pool_index] != None:
            overprov = whole_row[tot_virt_cap_index] / whole_row[arr_cap_tot_index] * 100
            util = whole_row[tot_used_cap_index] / whole_row[arr_cap_tot_index] * 100
            V3 = whole_row[arr_phy_cap_tot_index]
            W3 = whole_row[arr_phy_cap_rem]
            L3 = whole_row[tot_virt_cap_index]
            N3 = whole_row[compr_index]
            point_x = (V3-W3)/V3*100
            point_y = (L3/(1+(N3-1)*0.7))/V3*100
            if whole_row[sysname_index] in blacklist:
                print('blacklisted:',whole_row[sysname_index])
            else:
                found_points_xy.append([float(util),float(overprov),str(whole_row[sysname_index])+'-'+str(whole_row[pool_index]),float(whole_row[arr_cap_tot_index])])

    print('Found systems:',len(found_points_xy))
    return found_points_xy

def get_blacklist():
    #print('Checking blacklist')
    blacklist = []
    try:
        blacklist_file = open("blacklist")
        for line in blacklist_file:
            if line != '\n':
                for word in line.split(','):
                    blacklist.append(word.strip())

    except IOError:
        print("File \'blacklist\' not accessible")
        return blacklist

    blacklist_file.close()
    print('blacklist:',blacklist)

    return blacklist


if __name__ == '__main__':
    print(get_arrays_xlsx('2020.04.21xp_space.xlsx'))
